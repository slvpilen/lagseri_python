"""   lagseriereultat-calculater   """
# open .xlsx files only, NOT .xls (older then 2010)


from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from math import log10
import os 


#  pip install openpyxl
#  pip install pillow   # for filer med bilder


# Colors for better looking results in cmd
class colors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

    GREEN_B = '\x1b[6;30;42m'
    RED_B = '\x1b[6;30;41m'
    YELLOW_B = '\x1b[5;30;43m'
    LIGHT_B = '\x1b[6;30;44m'



def read_excel_files(direction):
    folder = os.listdir(direction)
    excel_files = []
    for file in folder:
        if file.endswith(".xlsx"):
            excel_files.append(file)
    return excel_files


def check_dato_in_qualification(dato_cell, start_date, end_date):  # denne er trøbbel!!! feil med format dato fra excel OG sjekk om dato er minder større enn qualicperiode
    correctDate = None
    try:
        dato_cell = dato_cell[8:10] + "." + dato_cell[5:7] + "." + dato_cell[0:4]
        newDate = datetime(int(dato_cell[6:11]),int(dato_cell[3:5]),int(dato_cell[0:2]))
        start = datetime(int(start_date[6:11]),int(start_date[3:5]),int(start_date[0:2]))
        end = datetime(int(end_date[6:11]),int(end_date[3:5]),int(end_date[0:2]))   
        correctDate = True
    except ValueError:
        correctDate = False
    if correctDate and start <= newDate <= end:
            return True
    else:
        if not correctDate:
            return "unvalid dato in sheet"
        else:
            return "date not in qualification periode"


def check_sheet(wb, filename, start_date, end_date):
    ok_sheets = []
    bad_sheets =  []

    mal_nvf = ["kropp", "kat", "fødsels", "navn", "lag", "rykk", "støt"]
    mal_nvf_5kamp = "k a m p"
    sheets_names = wb.sheetnames

    for sheet in sheets_names:
        ws = wb[sheet]  
        added = 0
        nvf_sheet = True

        seventh_row = [str(ws[str(get_column_letter(char)) + "7"].value) for char in range(1, 13)]  # 1-21/ A-U
        seventh_row = ''.join(seventh_row).lower()

        for elem in mal_nvf:
            if elem not in seventh_row:
                nvf_sheet = False 


        if nvf_sheet:
            if mal_nvf_5kamp in str(ws['G2'].value).lower():
                dato_cell = str(ws['V5'].value)  # 5-kamp protokoll
            else:
                dato_cell = str(ws['R5'].value)  # vanlig protokoll

            dato_in_ok_periode = check_dato_in_qualification(dato_cell, start_date, end_date)
            if dato_in_ok_periode != True:
                bad_sheets.append([sheet, f"right format, but {dato_in_ok_periode}"])
                continue

            for rad in range(9, 30):
                row = [ws[str(get_column_letter(char)) + str(rad)].value for char in range(1, 3)]

                try:
                    if str(row[0]).replace(".", "").replace(",","").replace("+","").isdigit() and str(row[1]).replace(".", "").replace(",","").replace("+","").isdigit():
                        ok_sheets.append(sheet)
                        added = True
                        break
                except:
                    continue


            if not added:
                bad_sheets.append([sheet, "right format, but no data"])
                
        else:
            bad_sheets.append([sheet, "wrong formate"])

        """
    Printing info about sheets passed and not passed check
    """        
    print(f"{colors.UNDERLINE}This sheets in file '{filename}' is ok: {colors.ENDC}") 
    for sheet in ok_sheets:
        print(f"{colors.GREEN_B}{sheet[:20]}{colors.ENDC}")
        
    print(f"{colors.UNDERLINE} \nThis is the bad ones: {colors.ENDC}") 
    for sheet in bad_sheets:
        print(f"{colors.RED_B} {sheet[0][:20].ljust(25)}({sheet[1]}) {colors.ENDC}") 
        
    print("-"*46)
    return ok_sheets, bad_sheets


    

class lifter:
    def __init__(self, data):
        # if: gir vanlig, else for 5-kamp
        print(data)
 
        if data[7].replace("-","0").split(".",1)[0].isdigit():   # Vanlig protokoll
            self.bw = float(data[1])
            self.category = data[2]
            if "k" in self.category.lower():
                self.gender = "f" 
            else:
                self.gender = "m"
            self.born = data[3]
            self.name = data[5]
            self.club = data[6].lower()
            self.attempts = [int(elem.replace("-","0").split(".",1)[0]) for elem in data[7:13]]


        elif not data[7].replace("-","0").split(".",1)[0].isdigit():   # 5-kamp
            self.bw = float(data[1])
            self.category = data[2]
            if "k" in self.category.lower():
                self.gender = "f" 
            else:
                self.gender = "m"
            self.born = data[4]
            self.name = data[6]
            self.club = data[7].lower()
            self.attempts = [int(str(elem).replace("-","0").split(".",1)[0]) for elem in data[8:14]]
        
    def get_total(self):
        self.snatch = self.attempts[0:3]
        self.cnj = self.attempts[3:6]

        if len(self.snatch) == 0 or len(self.cnj) == 0:
            return False
        
        self.good_snatch = False
        self.good_cnj = False

        self.best_snatch = 0
        self.best_cnj = 0
        for self.elem in self.snatch:
            if self.elem >= 1:
                self.good_snatch = True
                self.best_snatch = float(self.elem)
        for self.elem in self.cnj:
            if self.elem >= 1:
                self.good_cnj = True
                self.best_cnj = float(self.elem)

        if not self.good_snatch or not self.good_cnj:
            return False

        return self.best_snatch + self.best_cnj

    def sinclair_point(self, men_poeng=False):  # dersom True i parameter: blir det herrepoeng, uansett kjønn!
        self.men_poeng = men_poeng
        if self.gender.lower() == "m" or self.men_poeng == True and self.get_total()!=False:
            poeng = self.get_total()*(10**(0.751945030*((log10(self.bw/175.508))**2)))
            return poeng

        elif self.gender.lower() == "f" and self.get_total()!=False:
            poeng = self.get_total()*(10**(0.783497476*((log10(self.bw/153.655))**2)))
            return poeng
        else:
            return -1  # No valid result (need 1 good attempt in snatch and cnj to get a total)!
                   



def every_result(filename, direction, start_date, end_date, club, men_lagseri, women_lagseri):
    
    """
    open workbook and find sheets
    """
    wb = load_workbook(filename)  # her må filen være i aktuell mappe, legg til direction slik at åpner folder
    sheets_names = wb.sheetnames

    ok_sheets, bad_sheets = check_sheet(wb, filename, start_date, end_date)

    for sheet in ok_sheets:
        ws = wb[sheet]

        row = 9
        while True: 
            
            data = [str(ws[str(get_column_letter(char)) + str(row)].value) for char in range(1, 15)]
            if data.count('None') <= 1:

                lifter1 = lifter(data)

                if lifter1.club == club and lifter1.get_total() != False:
                    #print(lifter1.gender, lifter1.name, lifter1.club)
                    if lifter1.name not in men_lagseri:
                        men_lagseri[lifter1.name] = lifter1.sinclair_point(True)
                    else:
                        sinclaire = lifter1.sinclair_point(True)
                        if sinclaire > men_lagseri[lifter1.name]:
                            men_lagseri.update({lifter1.name: lifter1.sinclair_point(True)})


                    if lifter1.gender == "f":
                        if lifter1.name not in women_lagseri:
                            women_lagseri[lifter1.name] = lifter1.sinclair_point()
                        else:
                            sinclaire = lifter1.sinclair_point()
                            if sinclaire > women_lagseri[lifter1.name]:
                                women_lagseri.update({lifter1.name: lifter1.sinclair_point()})

  

            if ws['A' + str(row)].value == "Stevnets leder:" or row > 500:
                break 

            row+= 1


def info_from_user():
    def check_date(spm):
        svar = input(spm)
        if len(svar) != 10:
            print(f"{svar} is not a valid input. Try again! 1")
            check_date(spm)
        elif len(svar.replace(".","")) != 8:
            print(f"{svar} is not a valid input. Try again! 2")
            check_date(spm)
        elif svar[0:2].isdigit and svar[3:5].isdigit and svar[6:9].isdigit:
            # Check if input date is in the past:
            correctDate = None
            try:
                newDate = datetime(int(svar[6:11]),int(svar[3:5]),int(svar[0:2]))
                correctDate = True
            except ValueError:
                correctDate = False
            present = datetime.now()
            if correctDate and newDate <= present:
                return svar
            elif correctDate and newDate > present:
                print(f"{svar} is in the future. If you want to calculato until today, use todays dato!")
                check_date(spm)
            else:
                print(f"{svar} is not a valid input. Try again! 2")
                check_date(spm)
        else:
            print(f"{svar} is not a valid input. Try again! 3")
            check_date(spm)
            

    start_date = check_date("Write start date (dd.mm.yyyy): ")  # "01.01.2010"
    end_date = check_date("Write end date (dd.mm.yy): ")  # "01.12.2022"
    club = input("Enter the team/club you want to check. This need to be correct, no check!: ")  # "Nidelv IL" 

    print("-"*46)
    return start_date, end_date, club.lower()



def main():
    # Intro text:
    print("This is a program that calculate team results for a periode","\n", "-"*46)

    # getting nessesary info from user:
    start_date, end_date, club = info_from_user()

    # Getting all the excel files in direction (this folder)
    direction = os.path.dirname(os.path.realpath(__file__)) #r"C:\Users\oskar\Documents\lagserie_python"
    excel_files = read_excel_files(direction)


    men_lagseri = {}
    women_lagseri = {}

    # getting all the best women and men results:
    for fn in excel_files:
        print(f"{colors.OKBLUE}filename: {fn}{colors.ENDC}") 
        every_result(fn, direction, start_date, end_date, club, men_lagseri, women_lagseri)
    

    # Adding the results to an list and sort the lisr
    men_lagseri_liste =[]
    women_lagseri_liste =[]
    for navn in men_lagseri:
        men_lagseri_liste.append([navn, men_lagseri[navn]])
    men_lagseri_liste = sorted(men_lagseri_liste,key=lambda l:l[1], reverse=True)

    for navn in women_lagseri:
        women_lagseri_liste.append([navn, women_lagseri[navn]])
    women_lagseri_liste = sorted(women_lagseri_liste,key=lambda l:l[1], reverse=True)



    print("Mens sorted sinclair points:")
    for lifter in men_lagseri_liste:
        print(f"{colors.YELLOW_B}{lifter[0][:20].ljust(25)}\t{colors.ENDC}", f"{colors.LIGHT_B}{round(lifter[1],2)}{colors.ENDC}")

    print("\nWomens sorted sinclair points:")
    for lifter in women_lagseri_liste:
        print(f"{colors.YELLOW_B}{lifter[0][:20].ljust(25)}\t{colors.ENDC}", f"{colors.LIGHT_B}{round(lifter[1],2)}{colors.ENDC}")
    
    


if __name__ == "__main__":
    main()

# idè: ENdre slik at alle resultatet lagres i lifter objektet og lagre alle 
#løftere i en dictionery. Der kan man enkelt lage metoder i class som
# henter ut beste sinclair, guttepoeng/herrepoeng
# lage lister kun dersom alder er innenfor riktig etc